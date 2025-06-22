from openpyxl import load_workbook, Workbook
from captcha import patcha
from time import sleep
from bs4 import BeautifulSoup
import requests
from tqdm import tqdm
# import winsound
import os

# Путь к файлу
output_path = r"G:\Мой диск\ПР лизинг\База клиентов\Email\List_email.xlsx"  

wb = load_workbook(output_path)
ws = wb.active

headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
}

# Открываем файл с адресами
workbook = load_workbook(r"G:\Мой диск\ПР лизинг\База клиентов\Email\Email из Интернета.xlsm", data_only=True)
sheet = workbook["Адреса List"]

# Цикл по строкам
for row in tqdm(range(2, sheet.max_row + 1)):
    try:
        url_start = sheet.cell(row=row, column=3).value
        inn = sheet.cell(row=row, column=2).value
        if not url_start:
            continue

        req = requests.get(url_start, headers=headers)
        sleep(1)
        soup = BeautifulSoup(req.text, "lxml")

        try:
            url = "https://www.list-org.com/"  + soup.find('div', class_="org_list").find("a").get("href")
        except:
            patcha('Home')  # Капча
            req = requests.get(url_start, headers=headers)
            sleep(1)
            soup = BeautifulSoup(req.text, "lxml")
            url = "https://www.list-org.com/"  + soup.find('div', class_="org_list").find("a").get("href")

        req2 = requests.get(url, headers=headers)
        sleep(1)
        soup2 = BeautifulSoup(req2.text, "lxml")

        pre_inn = soup2.title.text
        client_name = pre_inn[:pre_inn.find(",")].strip()

        try:
            email = soup2.find('a', class_="wwbw").get("href").replace('mailto:', "").replace(',', ";").strip().lower()
            if "tensor.ru" in email:
                email = "Нет"
        except:
            email = "Нет"

        # Добавляем строку в конец таблицы
        ws.append([client_name, inn, email])

        # Сохраняем после каждой строки (или можно сделать раз в N строк)
        wb.save(output_path)

    except Exception as e:
        print(f"Ошибка на строке {row}: {e}")
    finally:
        # Финальное сохранение
        wb.save(output_path)
       
# winsound.Beep(frequency=1500, duration=1000)
print("Готово")