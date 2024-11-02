from openpyxl import Workbook
import re
import openpyxl
from time import sleep
import pandas as pd
from bs4 import BeautifulSoup
import requests
from tqdm import tqdm   
 
headers = {
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
}

data_email = []

workbook = openpyxl.load_workbook(r"G:\Мой диск\ПР лизинг\База клиентов\Email\Адреса для скачивания.xlsm", data_only=True)
sheet = workbook["Адреса ВБЦ"]  
for row in range(2,3): # sheet.max_row+1
    sleep(0.01)
    url = sheet.cell(row=row, column=3).value
    # response = requests.get(url,headers=headers)
    # sleep(2)                                                         
    with open('Vbankcenter/test.html', "r", encoding="utf-8") as f:
        data = f.read()
    soup = BeautifulSoup(data, "lxml")
    filename = soup.title.text
    client_name = filename[0:filename.find(",")-1]
    ogrn = url.replace("https://vbankcenter.ru/contragent/", "")
    email = "-"
    try:
        email_links = soup.find_all(href=re.compile("mailto:"))
        for item in email_links:
            if item.text.strip() != "client@vbankcenter.ru":
                email = item.text.strip()                         
        print(email)
    except:    
        continue
    
    # data_email.append([client_name,ogrn,email])
    # # # Create a DataFrame from the list
    # df_email = pd.DataFrame(data_email, columns =['client_name', 'ogrn', 'email'])
    # df_email.to_excel('SBIS download/Sbis_email.xlsx', index=True)


    # Если не работает, то добавить sleep 12 секунд





    