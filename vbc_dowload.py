from openpyxl import Workbook
import re
import openpyxl
from time import sleep
import pandas as pd
from bs4 import BeautifulSoup
import requests
from tqdm import tqdm
import winsound  
 
headers = {
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
}

data_email = []
workbook = openpyxl.load_workbook("G:\Мой диск\ПР лизинг\База клиентов\Email\Адреса для скачивания.xlsm", data_only=True)
sheet = workbook["Адреса ВБЦ"]  
сounter = 0
for row in tqdm(range(2,sheet.max_row+1)): # sheet.max_row+1
    sleep(0.01)
    сounter += 1
    if сounter == 150:
        winsound.Beep(frequency=1500,duration=1000)
        print("Нужно сменить IP")  
        input("Продолжить? Нажмите ENTER")
        counter=1  
    url = sheet.cell(row=row, column=3).value
    try:
        response = requests.get(url,headers=headers)
        sleep(9)     # если меньше 9 секунд - дает только 30 записей
        soup = BeautifulSoup(response.text, "lxml")
    except:
        winsound.Beep(frequency=1500,duration=1000)
        print("Нужно сменить IP")  
        input("Продолжить? Нажмите ENTER")
        counter =1  
        response = requests.get(url,headers=headers)
        soup = BeautifulSoup(response.text, "lxml")
    filename = soup.title.text
    client_name = filename[0:filename.find(",")-1]
    ogrn = url.replace("https://vbankcenter.ru/contragent/", "")
    email = "-"
    try:
        email_links = soup.find_all(href=re.compile("mailto:"))
        for item in email_links:
            if item.text.strip() != "client@vbankcenter.ru":
                if "tensor.ru" in item.text:
                    email = "-"
                else:
                    email = item.text.strip()                         
    except:    
        continue
    
    data_email.append([client_name,ogrn,email])
    df_email = pd.DataFrame(data_email)
    df_email.to_excel('g:\Мой диск\ПР лизинг\База клиентов\Email\VBC_email.xlsx', index=False)

winsound.Beep(frequency=1500,duration=1000)   # Если не работает, то добавить sleep 12 секунд





    