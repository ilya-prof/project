import email
from time import sleep
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
import requests
from tqdm import tqdm

headers = {
    "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
}
data_email = []

workbook = openpyxl.load_workbook(r"G:\Мой диск\ПР лизинг\База клиентов\Email\Адреса для скачивания.xlsm", data_only=True)
sheet = workbook["Адреса СБИС"]
for row in tqdm(range(2,sheet.max_row+1)): # sheet.max_row+1
    sleep(0.01) 
    url = sheet.cell(row,3).value
    response = requests.get(url,headers=headers)
    sleep(2)                                                         
    soup = BeautifulSoup(response.text, "lxml")
    filename = soup.title.text
    client_name = filename[0:filename.find("ИНН")-1]
    inn = url.replace("https://saby.ru/profile/", "")
    email = "-"
    try:
        email_links = soup.find_all(class_="contractorCard-ContactItem__text")
        for item in email_links:
            if "@" in item.text:
                email = item.text.strip()
                if "tensor.ru" in email:
                    email = "-"
    except:
        continue
    
    data_email.append([client_name,inn,email])
    # # Create a DataFrame from the list
    df_email = pd.DataFrame(data_email, columns =['client_name', 'inn', 'email'])
    df_email.to_excel(r"g:\Мой диск\ПР лизинг\База клиентов\Email\Sbis_email.xlsx", index=False)


    # Если не работает, то добавить sleep 12 секунд
