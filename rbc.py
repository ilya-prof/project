from time import sleep
import pandas as pd
from bs4 import BeautifulSoup
import requests
import openpyxl
from tqdm import tqdm
import time

data_list = []

headers = {
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
}
# Делаем цикл:
workbook = openpyxl.load_workbook(r"G:\Мой диск\ПР лизинг\База клиентов\Email\Адреса для скачивания.xlsm", data_only=True)
sheet = workbook["Адреса РБК"] 
for row in tqdm(range(2,sheet.max_row+1)): #sheet.max_row+1)  
    time.sleep(0.01) 
    url_start = sheet.cell(row=row, column =3).value
    
    # Ищем ссылку на страницу компании
    try:
        req = requests.get(url_start, headers=headers)
        soup=BeautifulSoup(req.text,"lxml")
        sleep(2)
    except:
        print("Необходимо изменить IP!!!")
        print('Продолжить?')
        req = requests.get(url_start, headers=headers)
        soup=BeautifulSoup(req.text,"lxml")
    try:
        url = soup.find("a", class_="company-name-highlight").get("href")
        sleep(2)
    except Exception as err:
        print(err)
        continue       
        
    # Проходим на страницу компании
    try:
        req2 = requests.get(url, headers=headers)
        sleep(2)
        soup2 = BeautifulSoup(req2.text,"lxml")
    except Exception as err:
        print(err)
    
    #Находим ИНН, ОГРН и название
    pre_inn = soup2.title.text
    inn = pre_inn[pre_inn.find("ИНН ")+4:pre_inn.find(" —")].strip()
    name = pre_inn[:pre_inn.find(" — ")].strip() 
    ogrn = pre_inn[pre_inn.find("ОГРН")+5:pre_inn.find(",")].strip()
    # Находим email 
    try:
        email = soup2.find(string="E-mail").next_element.text.strip().lower()
        if "tensor.ru" in email:
            email = "Нет"
    except:
        email = "Нет"
    
    # Все заводим в список
    data_list.append([name,ogrn,inn,email])
    df_email = pd.DataFrame(data_list, columns =['name','ogrn','inn','email'])
    df_email.to_excel(r"g:\Мой диск\ПР лизинг\База клиентов\Email\RBC_email.xlsx", index=False)
