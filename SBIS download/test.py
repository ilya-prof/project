import openpyxl
 
workbook = openpyxl.load_workbook(r"G:\Мой диск\ПР лизинг\База клиентов\Email\Адреса для скачивания.xlsm", data_only=True)
sheet = workbook["Адреса СБИС"]  
for row in range(1, sheet.max_row+1):
    url = sheet.cell(row=row, column=3).value
    print('url',url)



    